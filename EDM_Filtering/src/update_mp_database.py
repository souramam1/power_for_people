#!/usr/bin/env python3
"""
update_mp_database.py
=====================

What this script does
---------------------
Takes any Excel (.xlsx) workbook that contains a list of MPs, and adds (or
overwrites) a column with a yes/no flag pulled from edm_signing_status.csv
saying whether each MP has signed an EDM matching the filters used when
the status CSV was produced.

The original workbook is NEVER modified -- output is written to a new,
auto-dated copy alongside the source file (or to a custom path if you set
OUTPUT_XLSX in the CONFIG block).

Who this is for
---------------
You do NOT need to know Python. Open this file in a text editor, edit the
two main values in the CONFIG block (SOURCE_XLSX and NEW_COLUMN_HEADER),
save, then run:

    python update_mp_database.py

See README.md in this folder for setup.

Robustness
----------
The script validates every input before touching the spreadsheet, and
fails with a clear, human-readable message (no Python stack traces) if
anything is off:
  * source file missing / not a real .xlsx
  * paste looks like a web URL instead of a local path
  * sheet or name column can't be found (lists what WAS found)
  * status CSV or status column missing
  * new column header is empty

MP names are matched with a 3-tier strategy:
  1. Exact normalised "First Last" match (case/diacritic-insensitive).
  2. Token subset -- handles middle names ("Mary Foy" vs "Mary Kelly Foy").
  3. First-initial + surname overlap -- handles diminutives
     ("Peter" vs "Pete", "Edward" vs "Ed").

Cells filled via tier 2 or 3 are highlighted yellow so you can eyeball
them. Unmatched rows are left blank and listed on screen.
"""

from __future__ import annotations

import csv
import os
import re
import shutil
import sys
import unicodedata
from datetime import date
from pathlib import Path
from typing import NoReturn

# The script lives in <project>/src/. This is the project root -- relative
# paths in the CONFIG block below are resolved against it, so the script
# works whether you run it from the Code folder or from inside src/.
_PROJECT_ROOT = Path(__file__).resolve().parent.parent


# =============================================================================
# CONFIG -- edit these values, then run: python update_mp_database.py
# Everything below the "END OF CONFIG" line is internal logic.
# =============================================================================

# --- Input workbook -------------------------------------------------------
# Paste the path to the Excel file you want to update. Works with:
#   * absolute paths:  C:\Users\DELL\OneDrive\...\My Database.xlsx
#   * relative paths (to the project root):  databases/original_db/My DB.xlsx
#   * paths copied from File Explorer (quotes and backslashes are fine)
#
# IMPORTANT: paste a LOCAL FILE PATH, not a OneDrive "Share" web URL
# (https://onedrive.live.com/...). OneDrive files sync to your PC; open
# File Explorer, right-click the file, pick "Copy as path", and paste THAT.
SOURCE_XLSX = r"databases/original_db/MPs and Peers Database.xlsx"

# --- New column -----------------------------------------------------------
# The header for the column that will be created (or overwritten if it
# already exists) in the output workbook.
NEW_COLUMN_HEADER = "TEST COLUMN HEADER TO APPEND"

# --- Which yes/no value to use from the status CSV ------------------------
# STATUS_CSV: the signing-status CSV produced by EDM_Signatory_General_Check.py.
#   Default "edm_signing_status.csv" is the unfiltered master.
#   Point at a filtered run's CSV to use its yes/no, e.g.
#   "edm_signing_status_party-Labour_kw-climate.csv".
#
# STATUS_VALUE_COLUMN: which column of that CSV supplies the yes/no. Common:
#   "signed_edm_since_2024_07"              -- all EDMs in the window
#   "signed_edm_non_prayer_since_2024_07"   -- excluding prayers (default)
#   "signed_matching_edm"                   -- use this for a filtered run
STATUS_CSV          = "databases/edm_data/edm_signing_status.csv"
STATUS_VALUE_COLUMN = "signed_edm_non_prayer_since_2024_07"

# --- Output path ----------------------------------------------------------
# None = auto-generate a dated file inside databases/updated_db/ if that
# folder exists, otherwise alongside the source. Set to an explicit string
# (absolute or relative to project root) to override.
OUTPUT_XLSX = None

# --- Sheet / name-column hints (advanced) ---------------------------------
# Leave as None for auto-detection. If your workbook has multiple MP sheets
# or unusual column headers, set these explicitly to skip the guessing.
SHEET_NAME         = "MPs"  # e.g. "MPs"
FIRST_NAME_COLUMN  = None  # e.g. "First Name"
LAST_NAME_COLUMN   = None  # e.g. "Last Name"
FULL_NAME_COLUMN   = None  # e.g. "Name" (used when First/Last aren't present)

# =============================================================================
# END OF CONFIG -- do not edit below this line unless you know what you're doing.
# =============================================================================


def _abs(p: str) -> str:
    """Resolve a path against the project root. Absolute paths pass through."""
    path = Path(p)
    return str(path if path.is_absolute() else _PROJECT_ROOT / path)


# Resolve path-like CONFIG values once, at module load, so the rest of the
# code can treat them as plain absolute strings. SOURCE_XLSX and OUTPUT_XLSX
# are normalised separately inside validate_inputs() because they need quote
# stripping and URL rejection first.
STATUS_CSV = _abs(STATUS_CSV)


# ---------------------------------------------------------------------------
# Friendly error handling
# ---------------------------------------------------------------------------
def die(msg: str) -> NoReturn:
    """Print a clear error and exit -- no stack trace."""
    print(f"\nERROR: {msg}\n", file=sys.stderr)
    sys.exit(1)


# ---------------------------------------------------------------------------
# Input validation / normalisation
# ---------------------------------------------------------------------------
def normalise_source_path(raw: str) -> str:
    """Clean up whatever path-like string the user pasted into SOURCE_XLSX.
    Resolves relative paths against the project root."""
    if not raw or not raw.strip():
        die("SOURCE_XLSX is empty. Paste the path to your Excel file into "
            "the CONFIG block.")
    p = raw.strip()
    # Strip surrounding quotes (File Explorer's "Copy as path" wraps in ").
    if (p.startswith('"') and p.endswith('"')) or (p.startswith("'") and p.endswith("'")):
        p = p[1:-1]
    if p.lower().startswith(("http://", "https://")):
        die(f"SOURCE_XLSX looks like a web URL:\n  {p}\n"
            f"This script needs a LOCAL FILE path. In File Explorer, "
            f"right-click your file and choose \"Copy as path\", then paste "
            f"that into SOURCE_XLSX.")
    # Expand ~ so e.g. "~/Documents/..." works.
    p = os.path.expanduser(p)
    # Resolve relative paths against the project root.
    p = _abs(p)
    return p


def default_output_path(source_path: str) -> str:
    """Auto-generate a dated output path. Prefers databases/updated_db/ if
    that folder exists, else a dated sibling of the source. Preserves any
    existing _UPDATED_<dd>_<mm> suffix convention (replaces it with today)."""
    source = Path(source_path)
    base, ext = source.stem, source.suffix
    today = date.today().strftime("%d_%m")
    m = re.match(r"(.+)_UPDATED_\d+_\d+$", base)
    if m:
        base = m.group(1)
    new_name = f"{base}_UPDATED_{today}{ext}"
    updated_db = _PROJECT_ROOT / "databases" / "updated_db"
    out_dir = updated_db if updated_db.is_dir() else source.parent
    return str(out_dir / new_name)


def validate_inputs() -> tuple[str, str]:
    """Check every user-supplied input early. Returns (source_path, output_path)."""
    if not NEW_COLUMN_HEADER or not NEW_COLUMN_HEADER.strip():
        die("NEW_COLUMN_HEADER is empty. Set it to the header you want for "
            "the new column in the CONFIG block.")

    source = normalise_source_path(SOURCE_XLSX)
    if not os.path.exists(source):
        die(f"Source workbook not found:\n  {source}\n"
            f"Check that SOURCE_XLSX points to a file that exists on your "
            f"computer. If the file is on OneDrive, make sure it's synced "
            f"(not just \"Files on Demand\" / cloud-only).")
    if not source.lower().endswith((".xlsx", ".xlsm")):
        die(f"Source workbook must be a .xlsx (or .xlsm) file. Got:\n  {source}\n"
            f"If your file is .xls, open it in Excel and \"Save As\" .xlsx first.")

    if not os.path.exists(STATUS_CSV):
        die(f"Status CSV not found:\n  {STATUS_CSV}\n"
            f"Run EDM_Signatory_General_Check.py first to produce it, or point "
            f"STATUS_CSV at an existing file.")

    if OUTPUT_XLSX:
        output = _abs(OUTPUT_XLSX.strip())
    else:
        output = default_output_path(source)
    if os.path.abspath(output) == os.path.abspath(source):
        die(f"OUTPUT_XLSX points to the same file as SOURCE_XLSX:\n  {output}\n"
            f"The script never modifies the source -- choose a different "
            f"output path, or set OUTPUT_XLSX = None to auto-generate one.")
    out_dir = os.path.dirname(output) or "."
    os.makedirs(out_dir, exist_ok=True)

    return source, output


# ---------------------------------------------------------------------------
# Name normalisation & fuzzy matching
# ---------------------------------------------------------------------------
def norm(s: str | None) -> str:
    """Lower-case, strip diacritics, collapse whitespace, normalise hyphens
    and apostrophes. Returns "" for None/empty."""
    if not s:
        return ""
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    # Normalise curly quotes / apostrophes to ASCII.
    s = s.replace("’", "'").replace("‘", "'")
    # Hyphens -> spaces so "Long-Bailey" matches "Long Bailey" and
    # "Lewell-Buck" yields {lewell, buck} tokens for subset matching.
    s = s.replace("-", " ")
    return " ".join(s.lower().split())


def tokens(s: str) -> frozenset[str]:
    return frozenset(norm(s).split())


def split_full_name(full: str) -> tuple[str, str]:
    """Heuristically split a single-cell full name into (first, last).
    First token -> first name; everything else -> last name."""
    parts = norm(full).split()
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], " ".join(parts[1:])


# ---------------------------------------------------------------------------
# Load status CSV
# ---------------------------------------------------------------------------
def load_status() -> list[dict]:
    """Return each MP row from the status CSV as:
        {"full": normalised full name, "tokens": frozenset, "flag": "yes"/"no"}.
    Raises (friendly error) if STATUS_VALUE_COLUMN isn't in the CSV."""
    rows: list[dict] = []
    with open(STATUS_CSV, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if STATUS_VALUE_COLUMN not in (reader.fieldnames or []):
            die(f"Status CSV {STATUS_CSV!r} doesn't have a column named "
                f"{STATUS_VALUE_COLUMN!r}.\nAvailable columns: "
                f"{reader.fieldnames}\nFix STATUS_VALUE_COLUMN in the CONFIG "
                f"block (common choices: \"signed_edm_since_2024_07\", "
                f"\"signed_edm_non_prayer_since_2024_07\", "
                f"\"signed_matching_edm\").")
        for row in reader:
            name = (row.get("name") or "").strip()
            flag = (row.get(STATUS_VALUE_COLUMN) or "").strip().lower()
            if not name or flag not in ("yes", "no"):
                continue
            full = norm(name)
            rows.append({"full": full,
                          "tokens": frozenset(full.split()),
                          "flag": flag})
    if not rows:
        die(f"Status CSV {STATUS_CSV!r} was loaded but yielded 0 usable rows "
            f"(expected yes/no in column {STATUS_VALUE_COLUMN!r}).")
    return rows


def find_match(first: str, last: str, mps: list[dict]) -> tuple[str | None, str]:
    """Return (flag, mode) where mode is 'exact', 'fuzzy', or 'none'."""
    xlsx_full = norm(f"{first or ''} {last or ''}")
    xlsx_toks = frozenset(xlsx_full.split())
    xlsx_last_toks = frozenset(norm(last or "").split())
    xlsx_first_initial = norm(first or "")[:1]

    if not xlsx_toks:
        return None, "none"

    # 1. Exact full-name match.
    for m in mps:
        if m["full"] == xlsx_full:
            return m["flag"], "exact"

    def surname_overlap(m):
        return bool(xlsx_last_toks & m["tokens"])

    # 2. Token subset -- handles middle names.
    subset_hits = [
        m for m in mps
        if surname_overlap(m)
        and (xlsx_toks <= m["tokens"] or m["tokens"] <= xlsx_toks)
    ]
    if len(subset_hits) == 1:
        return subset_hits[0]["flag"], "fuzzy"

    # 3. First-initial + surname overlap -- handles diminutives.
    initial_hits = [
        m for m in mps
        if surname_overlap(m)
        and any(
            t.startswith(xlsx_first_initial)
            for t in m["tokens"] - xlsx_last_toks
        )
    ]
    if len(initial_hits) == 1:
        return initial_hits[0]["flag"], "fuzzy"

    return None, "none"


# ---------------------------------------------------------------------------
# Sheet / column detection
# ---------------------------------------------------------------------------
SHEET_CANDIDATES = ["MPs", "MP", "Members", "Commons", "MPs and Peers"]
FIRST_NAME_CANDIDATES = ["First Name", "Firstname", "First", "Given Name"]
LAST_NAME_CANDIDATES  = ["Last Name", "Lastname", "Last", "Surname", "Family Name"]
FULL_NAME_CANDIDATES  = ["Name", "MP Name", "Member Name", "Full Name",
                         "Display Name", "MP"]


def find_header(headers: dict[str, int], candidates: list[str]) -> int | None:
    """Return the 1-based column index of the first matching header, else None."""
    lower = {h.lower(): idx for h, idx in headers.items()}
    for c in candidates:
        if c.lower() in lower:
            return lower[c.lower()]
    return None


def detect_sheet(wb, hint: str | None):
    """Return the worksheet to use."""
    if hint:
        if hint not in wb.sheetnames:
            die(f"Sheet {hint!r} not found in workbook.\n"
                f"Sheets in workbook: {wb.sheetnames}\n"
                f"Fix SHEET_NAME in the CONFIG block.")
        return wb[hint]

    lower_map = {s.lower(): s for s in wb.sheetnames}
    for c in SHEET_CANDIDATES:
        if c.lower() in lower_map:
            return wb[lower_map[c.lower()]]

    # Fall back: pick the first sheet that looks like it has MP-style columns.
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = read_headers(ws)
        if find_header(headers, FIRST_NAME_CANDIDATES) \
           or find_header(headers, FULL_NAME_CANDIDATES):
            return ws

    die(f"Couldn't auto-detect an MP sheet in workbook.\n"
        f"Sheets found: {wb.sheetnames}\n"
        f"Set SHEET_NAME in the CONFIG block to the right one.")


def read_headers(ws) -> dict[str, int]:
    """Return {header_text: 1-based column index} from row 1 of the worksheet."""
    return {str(c.value).strip(): c.column
            for c in ws[1]
            if c.value is not None and str(c.value).strip()}


def detect_name_columns(ws, headers: dict[str, int]) -> tuple[str, int, int | None]:
    """Work out how the worksheet stores MP names.

    Returns (mode, primary_col, secondary_col) where:
      * mode='split': primary_col=First Name, secondary_col=Last Name
      * mode='full':  primary_col=Full Name,  secondary_col=None
    """
    # Explicit hints win outright.
    if FIRST_NAME_COLUMN and LAST_NAME_COLUMN:
        if FIRST_NAME_COLUMN not in headers:
            die(f"FIRST_NAME_COLUMN {FIRST_NAME_COLUMN!r} not found in sheet "
                f"{ws.title!r}.\nHeaders in sheet: {list(headers)}")
        if LAST_NAME_COLUMN not in headers:
            die(f"LAST_NAME_COLUMN {LAST_NAME_COLUMN!r} not found in sheet "
                f"{ws.title!r}.\nHeaders in sheet: {list(headers)}")
        return ("split", headers[FIRST_NAME_COLUMN], headers[LAST_NAME_COLUMN])
    if FULL_NAME_COLUMN:
        if FULL_NAME_COLUMN not in headers:
            die(f"FULL_NAME_COLUMN {FULL_NAME_COLUMN!r} not found in sheet "
                f"{ws.title!r}.\nHeaders in sheet: {list(headers)}")
        return ("full", headers[FULL_NAME_COLUMN], None)

    # Auto-detect.
    first = find_header(headers, FIRST_NAME_CANDIDATES)
    last  = find_header(headers, LAST_NAME_CANDIDATES)
    if first and last:
        return ("split", first, last)
    full  = find_header(headers, FULL_NAME_CANDIDATES)
    if full:
        return ("full", full, None)

    die(f"Couldn't find an MP name column in sheet {ws.title!r}.\n"
        f"Headers in sheet: {list(headers)}\n"
        f"Set FIRST_NAME_COLUMN + LAST_NAME_COLUMN (or FULL_NAME_COLUMN) "
        f"in the CONFIG block.")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def describe_config(source: str, output: str) -> None:
    print("--- Run configuration -------------------------------------------",
          file=sys.stderr)
    print(f"  Source workbook:     {source}", file=sys.stderr)
    print(f"  Output workbook:     {output}", file=sys.stderr)
    print(f"  New column header:   {NEW_COLUMN_HEADER}", file=sys.stderr)
    print(f"  Status CSV:          {STATUS_CSV}", file=sys.stderr)
    print(f"  Status value column: {STATUS_VALUE_COLUMN}", file=sys.stderr)
    print(f"  Sheet hint:          {SHEET_NAME or '(auto-detect)'}", file=sys.stderr)
    print(f"  Name column hints:   "
          f"first={FIRST_NAME_COLUMN or '(auto)'}, "
          f"last={LAST_NAME_COLUMN or '(auto)'}, "
          f"full={FULL_NAME_COLUMN or '(auto)'}", file=sys.stderr)
    print("-----------------------------------------------------------------",
          file=sys.stderr)


def main() -> None:
    # openpyxl import is guarded so a missing install gives a friendly error
    # instead of an ImportError stack trace.
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
    except ImportError:
        die("The 'openpyxl' library isn't installed. From this folder, run:\n"
            "  python -m pip install -r requirements.txt")

    source, output = validate_inputs()
    describe_config(source, output)

    mps = load_status()
    print(f"Loaded {len(mps)} MPs from {STATUS_CSV}", file=sys.stderr)

    # Copy the source to the output path BEFORE opening -- openpyxl rewrites
    # the file when saved, so we never want to overwrite SOURCE_XLSX.
    try:
        shutil.copyfile(source, output)
    except OSError as e:
        die(f"Couldn't copy source to output:\n  {source}\n  -> {output}\n"
            f"Reason: {e}\nIs the output file open in Excel? Close it and retry.")
    print(f"Copied source -> {output}", file=sys.stderr)

    try:
        wb = load_workbook(output)
    except Exception as e:
        die(f"Couldn't open {output!r} as an Excel workbook.\nReason: {e}\n"
            f"Check that SOURCE_XLSX is a real .xlsx file and not corrupted.")

    ws = detect_sheet(wb, SHEET_NAME)
    headers = read_headers(ws)
    mode, col_a, col_b = detect_name_columns(ws, headers)

    print(f"Using sheet {ws.title!r}. Name mode: {mode} "
          f"(columns {col_a}{'+'+str(col_b) if col_b else ''}).", file=sys.stderr)

    # Locate / create the target column.
    if NEW_COLUMN_HEADER in headers:
        target_col = headers[NEW_COLUMN_HEADER]
        print(f"Column {NEW_COLUMN_HEADER!r} already exists at column "
              f"{target_col}; values will be overwritten.", file=sys.stderr)
    else:
        target_col = ws.max_column + 1
        ws.cell(row=1, column=target_col, value=NEW_COLUMN_HEADER)
        print(f"Added new column {NEW_COLUMN_HEADER!r} at column {target_col} "
              f"({ws.cell(row=1, column=target_col).coordinate}).",
              file=sys.stderr)

    fuzzy_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC",
                             fill_type="solid")

    exact, fuzzy, unmatched, blank = 0, [], [], 0
    for r in range(2, ws.max_row + 1):
        if mode == "split":
            first = ws.cell(row=r, column=col_a).value
            last  = ws.cell(row=r, column=col_b).value
        else:
            full = ws.cell(row=r, column=col_a).value
            first, last = split_full_name(str(full) if full is not None else "")
        if not (first or last):
            blank += 1
            continue

        flag, match_mode = find_match(str(first or ""), str(last or ""), mps)
        cell = ws.cell(row=r, column=target_col)
        if match_mode == "exact":
            cell.value = flag
            cell.fill = PatternFill(fill_type=None)  # clear any prior fuzzy highlight
            exact += 1
        elif match_mode == "fuzzy":
            cell.value = flag
            cell.fill = fuzzy_fill
            fuzzy.append((r, first, last, flag))
        else:
            cell.value = None
            cell.fill = PatternFill(fill_type=None)
            unmatched.append((r, first, last))

    try:
        wb.save(output)
    except OSError as e:
        die(f"Couldn't save {output!r}.\nReason: {e}\n"
            f"Is the file open in Excel? Close it and retry.")

    # Summary
    print("", file=sys.stderr)
    print(f"Exact matches : {exact}", file=sys.stderr)
    print(f"Fuzzy matches : {len(fuzzy)} "
          f"(highlighted yellow -- verify these)", file=sys.stderr)
    for r, f, l, flag in fuzzy:
        print(f"  row {r}: {f!r} {l!r} -> {flag}", file=sys.stderr)
    print(f"Unmatched     : {len(unmatched)} (left blank)", file=sys.stderr)
    for r, f, l in unmatched:
        print(f"  row {r}: {f!r} {l!r}", file=sys.stderr)
    if blank:
        print(f"Skipped blank rows: {blank}", file=sys.stderr)
    print(f"\nSaved {output}", file=sys.stderr)


if __name__ == "__main__":
    main()
